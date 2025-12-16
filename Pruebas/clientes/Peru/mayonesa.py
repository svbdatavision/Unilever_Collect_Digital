# =====================================================
# 0. Importación de librerías y módulos utilitarios
# =====================================================
import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import messagebox


# =====================================================
# 1. Localización dinámica de la carpeta raíz del proyecto
# =====================================================
def _project_root():
    if getattr(sys, "frozen", False):
        macos_dir = os.path.dirname(sys.executable)
        contents_dir = os.path.dirname(macos_dir)
        app_bundle = os.path.dirname(contents_dir)
        return os.path.dirname(app_bundle)

    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


# =====================================================
# 2. Función principal del proceso
# =====================================================
def procesar():

    root = _project_root()

    # Rutas dinámicas con país = Colombia
    rutas = {
        "remittance": os.path.join(root, "Archivos", "Remittance", "Peru", "Remittance_Mayorsa.xlsx"),
        "salida": os.path.join(root, "Archivos", "Template", "Peru", "Remittance_Mayorsa.xlsx")
    }

    try:
        # =====================================================
        # 3. Lectura de Remittance
        # =====================================================
        remittance_df = pd.read_excel(rutas["remittance"])
        remittance_df.columns = remittance_df.columns.str.strip()

        # Limpieza de campo Comprobante
        if "Comprobante" in remittance_df.columns:
            remittance_df["Comprobante"] = (
                remittance_df["Comprobante"]
                .str.replace(r'^(01|07|D1)-F', 'F', regex=True)
                .str.replace(r'^(01|D1)-1', '1', regex=True)
            )

        tipo_map = {
            "Nota Credito": "NC",
            "Factura": "Factura",
        }

        # Construcción de formato de salida
        formato_df = pd.DataFrame()
        formato_df["Tipo Doc"] = remittance_df["Tipo Comprobante"].map(
            lambda x: tipo_map.get(x, "Otro")
        )
        formato_df["Referencia / Factura"] = remittance_df["Comprobante"]

        # NC → monto negativo
        formato_df["Monto"] = remittance_df.apply(
            lambda row: -row["Monto"] if row["Tipo Comprobante"] == "Nota Credito" else row["Monto"],
            axis=1
        )

        # Reglas de razon de descuento
        def razon_descuento(row):
            ref = str(row["Comprobante"])
            if ref.startswith("100") or ref.startswith("FN"):
                return "657"
            if ref.startswith("F701") or ref.startswith("F121"):
                return ""
            return tipo_map.get(row["Tipo Comprobante"], "Otro")

        formato_df["Razon de Descuento"] = remittance_df.apply(razon_descuento, axis=1)

        # =====================================================
        # 4. Nombre de archivo de salida dinámico
        # =====================================================
        ruta_salida = rutas["salida"]

        contador = 1
        while os.path.exists(ruta_salida):
            ruta_salida = os.path.join(
                os.path.dirname(rutas["salida"]),
                f"Remittance_Mayorsa_{contador}.xlsx"
            )
            contador += 1

        # Si existe, reemplazar
        if os.path.exists(ruta_salida):
            os.remove(ruta_salida)

        # Exportar
        formato_df.to_excel(ruta_salida, index=False)

        # =====================================================
        # 5. Escribir datos extra en columnas F y G con estilo
        # =====================================================
        wb = load_workbook(ruta_salida)
        ws = wb.active

        datos_extra = [
            ("Nombre Cliente", "Mayorsa"),
            ("Numero de Cliente", "10263122"),
            ("Referencia de Pago", ""),
            ("Pago", sum(formato_df["Monto"])),
            ("Metodo de Pago", "TRANSFERENCIA"),
            ("Fecha de pago", ""),
        ]

        fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
        font_negrita = Font(bold=True)
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        for i, (col_f, col_g) in enumerate(datos_extra, start=1):
            celda_f = ws[f"F{i}"]
            celda_g = ws[f"G{i}"]

            celda_f.value = col_f
            celda_g.value = col_g

            celda_f.fill = fill_azul
            celda_f.font = font_negrita
            celda_f.border = borde_negro
            celda_g.border = borde_negro

        wb.save(ruta_salida)

        messagebox.showinfo("¡Éxito!", f"✅ Archivo exportado como {ruta_salida}")

    except Exception as e:
        # Esta línea DEBE permanecer según tu instrucción
        print(f"Ocurrió un error: {e}")
