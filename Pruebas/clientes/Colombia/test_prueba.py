# =====================================================
# 0. Importación de librerías y módulos utilitarios
# =====================================================
import sys  # Para detectar ejecución empaquetada (frozen) y resolver rutas
import os   # Para construir rutas relativas al proyecto
import pandas as pd  # Principal herramienta para manipulación tabular
import numpy as np   # Utilidades numéricas/condicionales (np.select, np.where)
from openpyxl import Workbook, load_workbook  # <-- Importamos Workbook también

ruta_archivo = r"C:\Pruebas python\Remittance_olimpica.xlsx"
ruta_template = r"C:\Pruebas python\template_olimpica.xlsx"

# Leer el archivo original

columnas_necesarias = ["Código de Documento", "No. Doc", "Total a Pagar"]
df = pd.read_excel(ruta_archivo, skiprows=21)

# Filtrar solo las columnas necesarias
df = df[columnas_necesarias]


# Si el template no existe, lo creamos
if not os.path.exists(ruta_template):
    wb = Workbook()
    wb.save(ruta_template)
    print("✅ Template creado porque no existía.")

# Abrimos el template
wb = load_workbook(ruta_template)
ws = wb.active

# Limpiar datos previos
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# Escribir encabezados
for col_num, col_name in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_num, value=col_name)

# Escribir datos
for row_num, row_data in enumerate(df.values, start=2):
    for col_num, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=value)

# Guardar cambios
wb.save(ruta_template)
print("✅ Datos copiados al template exitosamente.")

