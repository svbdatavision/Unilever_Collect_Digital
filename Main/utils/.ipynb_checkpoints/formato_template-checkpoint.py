import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from copy import copy
from openpyxl.utils import get_column_letter
from datetime import datetime

    # =====================================================
    # 16. Aplicación de formato y exportación del Template final 
    # =====================================================

def copiar_hoja(origen, destino, nombre="Remittance"):
    """
    Utilidad: Copia una hoja completa de Excel (valores + estilos + dimensiones + merges).
    """
    ws_new = destino.create_sheet(title=nombre)

    # Copiar celdas y estilos
    for row in origen.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copiar dimensiones de columnas y filas
    for col_letter, col_dim in origen.column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width
    for row_idx, row_dim in origen.row_dimensions.items():
        ws_new.row_dimensions[row_idx].height = row_dim.height

    # Copiar merges
    for merge in origen.merged_cells.ranges:
        ws_new.merge_cells(str(merge))

    return ws_new


def exportar_template(
    hrc_template,
    suma_remittance,
    numero_orden,
    id_cliente,
    nombre_cliente,
    ruta_remittance,
    ruta_salida
):

    # Ajustamos a formato Template

    # Ordenar el DataFrame antes de exportar

    # Columna auxiliar para el orden personalizado de "Motivo del descuento"
    def orden_motivo(valor):
        if pd.isna(valor) or str(valor).strip() == "":
            return 0  # primero los vacíos
        elif str(valor).strip().upper() in ["384", "WOB"]:
            return 2  # últimos
        else:
            return 1  # el resto en el medio

    hrc_template["_orden_motivo"] = hrc_template["Motivo del descuento"].apply(orden_motivo)

    # Ordenar: primero por la columna auxiliar (0 -> vacíos, 1 -> medio, 2 -> últimos),
    # y dentro de cada grupo ordenar "Motivo del descuento" en orden DESCENDENTE.
    hrc_template = hrc_template.sort_values(
        by=["_orden_motivo", "Descuento", "Tipo de Documento", "Referencia / Factura"],
        ascending=[True, True, True, True]   # _orden_motivo ascendente, Descuento ascendente
    ).reset_index(drop=True)

    # Eliminar columna auxiliar
    hrc_template = hrc_template.drop(columns=["_orden_motivo"])
    
    
    # Exportar DataFrame a Excel en hoja "Template"
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        hrc_template.to_excel(writer, index=False, sheet_name="Template", startrow=17, startcol=2)

    wb = load_workbook(ruta_salida)
    ws = wb["Template"]

    # Colores y estilos
    azul_oscuro = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    celeste_intenso = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    letra_blanca = Font(color="FFFFFF")
    letra_negra = Font(color="000000")

    # Encabezados dinámicos
    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = numero_orden
    ws["C6"] = "Cliente"
    ws["C8"] = "Codigo de Cliente"
    ws["D6"] = nombre_cliente
    ws["D8"] = id_cliente
    ws["C11"] = "Referencia"
    ws["C12"] = numero_orden
    ws["D11"] = "Fecha"
    ws["D12"] = datetime.now().strftime("%d/%m/%Y") # fecha de hoy
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = suma_remittance

    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = suma_remittance
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = suma_remittance
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = suma_remittance -suma_remittance

    # Formatos numéricos y de texto
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'
    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Colores en celdas clave
    for cell in ["G2", "C6", "C7", "C8", "F6", "F7", "F8", "C11", "D11", "E11",
                 "F11", "C18", "D18", "E18", "F18", "G18", "H18", "I18"]:
        ws[cell].fill = azul_oscuro
        ws[cell].font = letra_blanca
    for cell in ["C4", "D6", "D7", "D8", "G6", "G7", "G8"]:
        ws[cell].fill = celeste_intenso
        ws[cell].font = letra_blanca
    ws["H2"].fill = amarillo
    ws["H2"].font = letra_negra

    # Ajuste dinámico de ancho de columnas
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)) and cell.number_format:
                        val_str = f"{cell.value:,.2f}" if "0.00" in cell.number_format else str(cell.value)
                    else:
                        val_str = str(cell.value)
                    max_length = max(max_length, len(val_str))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Copiar hoja Remittance original al archivo final
    wb_rem = load_workbook(ruta_remittance, data_only=False)
    ws_original = wb_rem.active
    copiar_hoja(ws_original, wb, nombre="Remittance")

    # Guardar archivo final
    wb.save(ruta_salida)
    print(f"✅ Archivo exportado con formato: {ruta_salida}")