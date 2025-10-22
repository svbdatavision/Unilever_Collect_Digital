import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import messagebox

# --- PATRONES Y SINÓNIMOS ---
DOC_PATTERNS = [
    re.compile(r'(F701-\d{3,})'),
    re.compile(r'(F121-\d{3,})'),
    re.compile(r'(FA\d{2,}-\d{3,})'),
]
TIPO_KEYWORDS = [
    'FACTURA', 'NOTA DE CRÉDITO', 'NOTA DE CREDITO',
    'NOTA DE DÉBITO', 'NOTA DE DEBITO',
    'PAGO DETRACCIÓN', 'PAGO DETRACCION',
    'AUTO-DETRACCIÓN', 'AUTO-DETRACCION',
    'FACTURA DEUDOR METRO', 'PROVEEDOR', 'DEUDOR'
]
HEADER_SYNONYMS = {
    'doc': [
        'nro. documento', 'nro documento', 'numero documento', 'número documento',
        'num documento', 'documento', 'referencia', 'ref', 'nro'
    ],
    'tipo': [
        'tipo', 'tipo documento', 'clase', 'clase doc', 'document type'
    ],
    'importe': [
        'neto', 'importe', 'importe factura', 'monto', 'total', 'importe neto', 'valor'
    ],
    'evitar_importe': ['retencion', 'retención', 'moneda', 'fecha de pago', 'fecha']
}
TIPO_MAP = {
    'NOTA DE CRÉDITO PROVEEDOR': 'NC',
    'NOTA DE CREDITO PROVEEDOR': 'NC',
    'FACTURA PROVEEDOR': 'Factura',
    'Auto-Detracción Deud': 'Fact. Convenio',
    'Auto-Detraccion Deud': 'Fact. Convenio',
    'Pago detracción': 'Fact. Convenio',
    'Pago detraccion': 'Fact. Convenio',
    'FACTURA DEUDOR METRO': 'Fact. Convenio',
}

def _norm(s):
    if pd.isna(s):
        return ''
    s = str(s).strip().lower()
    t = str.maketrans('áéíóúäëïöü', 'aeiouaeiou')
    return s.translate(t)

def _extract_doc(value):
    s = '' if pd.isna(value) else str(value)
    s = re.sub(r'^(?:\d{2}|0\d)-(?=F)', '', s)
    for pat in DOC_PATTERNS:
        m = pat.search(s)
        if m:
            return m.group(1)
    return ''

def _score_doc(series, header):
    header_hit = any(alias in _norm(header) for alias in HEADER_SYNONYMS['doc'])
    values = series.dropna().astype(str)
    extracted = values.apply(_extract_doc)
    frac_match = (extracted != '').mean() if len(values) else 0.0
    only_digits = values.str.fullmatch(r'\d+').mean() if len(values) else 0.0
    return 0.60*frac_match + 0.25*header_hit - 0.10*only_digits

def _score_tipo(series, header):
    header_hit = any(alias in _norm(header) for alias in HEADER_SYNONYMS['tipo'])
    values = series.dropna().astype(str).str.upper()
    hits = 0
    for kw in TIPO_KEYWORDS:
        hits += values.str.contains(kw.upper(), regex=False).sum()
    frac_kw = hits / max(len(values), 1)
    return 0.50*frac_kw + 0.40*header_hit

def _score_importe(series, header, doc_mask=None, tipo_series=None):
    h = _norm(header)
    if any(bad in h for bad in HEADER_SYNONYMS['evitar_importe']):
        return -1.0
    header_hit = 0.0
    for i, key in enumerate(['neto', 'importe', 'total', 'monto', 'valor']):
        if key in h:
            header_hit = 0.30 - 0.05*i
            break
    s = pd.to_numeric(series, errors='coerce')
    numeric_frac = s.notna().mean()
    co_coverage = s.notna()[doc_mask].mean() if (doc_mask is not None and doc_mask.any()) else 0.0
    sign_score = 0.0
    if tipo_series is not None:
        tipos = tipo_series.astype(str).str.upper()
        nc_mask = tipos.str.contains('NOTA DE CR|NOTA DE CREDITO')
        fa_mask = tipos.str.contains('FACTURA')
        sample = s.copy()
        neg_ok = (sample < 0)[nc_mask].mean() if nc_mask.any() else 0.0
        pos_ok = (sample >= 0)[fa_mask].mean() if fa_mask.any() else 0.0
        sign_score = 0.20 * (neg_ok + pos_ok)
    return 0.40*numeric_frac + 0.40*co_coverage + header_hit + sign_score

def detectar_columnas(df):
    doc_scores = {c: _score_doc(df[c], c) for c in df.columns}
    doc_col = max(doc_scores, key=doc_scores.get)
    doc_extracted = df[doc_col].apply(_extract_doc)
    doc_mask = doc_extracted != ''
    tipo_scores = {c: _score_tipo(df[c], c) for c in df.columns}
    tipo_scores[doc_col] = -1
    tipo_col = max(tipo_scores, key=tipo_scores.get)
    tipo_col = tipo_col if tipo_scores[tipo_col] > 0 else None
    tipo_series = df[tipo_col] if tipo_col else None
    imp_scores = {}
    for c in df.columns:
        if c == doc_col or (tipo_col and c == tipo_col):
            continue
        imp_scores[c] = _score_importe(df[c], c, doc_mask=doc_mask, tipo_series=tipo_series)
    importe_col = max(imp_scores, key=imp_scores.get)
    return doc_col, tipo_col, importe_col, doc_extracted

def _inferir_tipo_por_doc(doc):
    if not doc:
        return 'Otro'
    if doc.startswith('F121-'):
        return 'NOTA DE CRÉDITO PROVEEDOR'
    if doc.startswith('F701-'):
        return 'FACTURA PROVEEDOR'
    if doc.startswith('FA'):
        return 'Pago detracción'
    return 'Otro'

def _map_tipo_resumido(tipo_raw):
    return TIPO_MAP.get(tipo_raw, 'Otro')

def _razon_descuento(ref, tipo_raw):
    ref = str(ref)
    if ref.startswith('FA') or ref.startswith('FN'):
        return '657'
    if ref.startswith('F701-') or ref.startswith('F121-'):
        return ''
    return TIPO_MAP.get(tipo_raw, 'Otro')

def transformar(archivo):
    df = pd.read_excel(archivo)
    df.columns = [c.strip() for c in df.columns]
    doc_col, tipo_col, importe_col, doc_extracted = detectar_columnas(df)
    if tipo_col:
        tipo_raw = df[tipo_col].fillna('Otro').astype(str)
    else:
        tipo_raw = doc_extracted.map(_inferir_tipo_por_doc)
    out = pd.DataFrame()
    out['Tipo Doc'] = tipo_raw.map(_map_tipo_resumido)
    out['Referencia / Factura'] = doc_extracted
    out['Importe de factura'] = pd.to_numeric(df[importe_col], errors='coerce')
    out = out[out['Referencia / Factura'].astype(str).str.len() > 0].reset_index(drop=True)
    out['Razon de Descuento'] = [
        _razon_descuento(ref, raw) for ref, raw in zip(out['Referencia / Factura'], tipo_raw)
    ]
    return out

def procesar(archivo_remittance, _):
    try:
        formato_df = transformar(archivo_remittance)
        nombre_base = os.path.splitext(os.path.basename(archivo_remittance))[0]
        ruta_salida = os.path.join(
            os.path.dirname(archivo_remittance),
            f"Remmitance_Cencosud.xlsx"
        )
        contador = 1
        while os.path.exists(ruta_salida):
            ruta_salida = os.path.join(
                os.path.dirname(archivo_remittance),
                f"Remmitance_Cencosud_{contador}.xlsx"
            )
            contador += 1
        if os.path.exists(ruta_salida):
            os.remove(ruta_salida)
        formato_df.to_excel(ruta_salida, index=False)
        wb = load_workbook(ruta_salida)
        ws = wb.active
        datos_extra = [
            ("Nombre Cliente", "CENCOSUD RETAIL PERU S A"),
            ("Numero de Cliente", "10262842"),
            ("Referencia de Pago", ""),
            ("Pago", formato_df["Importe de factura"].sum()),
            ("Metodo de Pago", "TRANSFERENCIA"),
            ("Fecha de pago", ""),
        ]
        fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
        font_negrita = Font(bold=True)
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        for i, (col_f, col_g) in enumerate(datos_extra, start=1):
            celda_f = ws[f"F{i}"]
            celda_g = ws[f"G{i}"]
            celda_f.value = col_f
            celda_g.value = col_g
            celda_f.fill = fill_azul
            celda_f.font = font_negrita
            celda_f.border = borde_negro
            celda_g.border = borde_negro
        wb.save(ruta_salida)
        messagebox.showinfo("¡Éxito!", f"✅ Archivo exportado como {ruta_salida}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")
        raise e