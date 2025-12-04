import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import messagebox

def procesar(archivo_remittance, _):
    try:
        remittance_df = pd.read_excel(archivo_remittance)
        remittance_df.columns = remittance_df.columns.str.strip()

        if "Comprobante" in remittance_df.columns:
            remittance_df["Comprobante"] = (
                remittance_df["Comprobante"]
                .str.replace(r'^(01|07|D1)-F', 'F', regex=True)
                .str.replace(r'^(01|D1)-1', '1', regex=True)
            )

        tipo_map = {
            "Nota Credito": "NC",
            "Factura": "Factura",
        }

        formato_df = pd.DataFrame()
        formato_df["Tipo Doc"] = remittance_df["Tipo Comprobante"].map(lambda x: tipo_map.get(x, "Otro"))
        formato_df["Referencia / Factura"] = remittance_df["Comprobante"]

        # Si es NC, el monto es negativo
        formato_df["Monto"] = remittance_df.apply(
            lambda row: -row["Monto"] if row["Tipo Comprobante"] == "Nota Credito" else row["Monto"],
            axis=1)

        def razon_descuento(row):
            ref = str(row["Comprobante"])
            if ref.startswith("100") or ref.startswith("FN"):
                return "657"
            if ref.startswith("F701") or ref.startswith("F121"):
                return ""
            return tipo_map.get(row["Tipo Comprobante"], "Otro")

        formato_df["Razon de Descuento"] = remittance_df.apply(razon_descuento, axis=1)

        nombre_base = os.path.splitext(os.path.basename(archivo_remittance))[0]
        ruta_salida = os.path.join(
            os.path.dirname(archivo_remittance),
            f"Remmitance_Mayorsa.xlsx"
        )

        contador = 1
        while os.path.exists(ruta_salida):
            ruta_salida = os.path.join(
                os.path.dirname(archivo_remittance),
                f"Remmitance_Mayorsa_{contador}.xlsx"
            )
            contador += 1


        if os.path.exists(ruta_salida):
            os.remove(ruta_salida)

        formato_df.to_excel(ruta_salida, index=False)

        wb = load_workbook(ruta_salida)
        ws = wb.active

        datos_extra = [
            ("Nombre Cliente", "Mayorsa"),
            ("Numero de Cliente", "10263122"),
            ("Referencia de Pago", ""),
            ("Pago", sum(formato_df["Monto"])),
            ("Metodo de Pago", "TRANSFERENCIA"),
            ("Fecha de pago", ""),
        ]

        fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
        font_negrita = Font(bold=True)
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        for i, (col_f, col_g) in enumerate(datos_extra, start=1):
            celda_f = ws[f"F{i}"]
            celda_g = ws[f"G{i}"]

            celda_f.value = col_f
            celda_g.value = col_g

            celda_f.fill = fill_azul
            celda_f.font = font_negrita
            celda_f.border = borde_negro
            celda_g.border = borde_negro

        wb.save(ruta_salida)
        messagebox.showinfo("¡Éxito!", f"✅ Archivo exportado como {ruta_salida}")

    except Exception as e:
            print(f"Ocurrió un error: {e}")  # ← esta línea es obligatoria
