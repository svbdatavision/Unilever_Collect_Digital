import sys
import os
import re
import pandas as pd
import numpy as np
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font


def _project_root():
    """
    Devuelve la carpeta raíz del proyecto:
    - Si corre dentro de un .app -> la carpeta que contiene el .app
    - Si corre como script -> la carpeta del archivo actual (../)
    """
    if getattr(sys, "frozen", False):
        macos_dir = os.path.dirname(sys.executable)      # .../MyApp.app/Contents/MacOS
        contents_dir = os.path.dirname(macos_dir)        # .../MyApp.app/Contents
        app_bundle = os.path.dirname(contents_dir)       # .../MyApp.app
        return os.path.dirname(app_bundle)               # carpeta que contiene el .app
    return os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def procesar():
    root = _project_root()

    # --- Rutas ---
    rutas = {
        "remittance_pdf": os.path.join(root, "Archivos", "Remittance", "Remittance_D1.pdf"),
        "fbl5n":          os.path.join(root, "Archivos", "Base_de_datos", "FBL5N_d1.xlsx"),
        "salida":         os.path.join(root, "Archivos", "Template", "Template_HRC_D1.xlsx"),
    }

    # --- Paso 1: Extraer facturas del PDF de Remittance ---
    factura_pattern = re.compile(
        r"RE\s+\d+\s+PMP\d+\s+\d{1,3}(?:\.\d{3})*\s+\d+\s+\d{1,3}(?:\.\d{3})*\s+\d+\s+\d{1,3}(?:\.\d{3})*"
    )
    facturas = []
    with fitz.open(rutas["remittance_pdf"]) as doc:
        for page in doc:
            text = page.get_text()
            for match in factura_pattern.findall(text):
                parts = match.split()
                if len(parts) == 8:
                    facturas.append({
                        "TD": parts[0],
                        "Doc.Interno": parts[1],
                        "Referencia / Factura": parts[2],
                        "Valor Bruto": parts[3].replace('.', ''),
                        "Retenciones": parts[4],
                        "IVA": parts[5].replace('.', ''),
                        "Desc/Rec": parts[6],
                        "Neto Pagado": parts[7].replace('.', '')
                    })

    remittance = pd.DataFrame(facturas)
    remittance["Importe de factura"] = remittance["Neto Pagado"].astype(float)
    remittance["Tipo de Documento"] = "Factura"

    # --- Paso 2: Leer FBL5N y preparar ---
    FBL5N = pd.read_excel(
        rutas["fbl5n"],
        sheet_name="Sheet1",
        usecols=["Document Type", "Reference", "Amount in local currency", "Reason code", "Document Number", "Text"]
    )
    FBL5N = FBL5N[(FBL5N["Document Type"] == "RV") | (FBL5N["Reason code"] == "NRO")]

    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    FBL5N["Referencia / Factura"] = np.where(
        FBL5N["Reason code"] == "NRO",
        FBL5N["Document Number"].astype("Int64").astype(str),
        FBL5N["Referencia / Factura"]
    )

    # --- Paso 3: Merge y diferencias ---
    hrc_template = pd.merge(remittance, FBL5N, on="Referencia / Factura", how="left")
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )

    dif = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()
    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": dif["Referencia / Factura"],
        "Importe de factura": dif["Diferencia"],
        "Pago Neto": "",
        "Descuento": "MENORES VALORES",
        "Motivo del descuento": dif["Diferencia"].apply(
            lambda x: "WOB" if -20000 < x < 0 else ("384" if 0 < x < 20000 else "987")
        )
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # --- Paso 4: Comentarios y campos finales ---
    hrc_template["Comentarios"] = np.where(
        hrc_template["Tipo de Documento"] == "Factura", "",
        np.where(
            hrc_template["Descuento"] == "MENORES VALORES",
            hrc_template["Descuento"],
            hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("")
        )
    )
    cond_1 = (hrc_template["Motivo del descuento"] == "987") & (hrc_template["Importe de factura"] < -20000)
    cond_2 = (hrc_template["Motivo del descuento"] == "987") & (hrc_template["Importe de factura"] > 20000)
    hrc_template.loc[cond_1, "Comentarios"] = "Myr Vlr Pagado " + hrc_template.loc[cond_1, "Referencia / Factura"].fillna("")
    hrc_template.loc[cond_2, "Comentarios"] = "Saldo FC " + hrc_template.loc[cond_2, "Referencia / Factura"].fillna("")

    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    nota_credito = FBL5N[FBL5N["Reason code"] == "NRO"].copy()
    nota_credito["Tipo de Documento"] = "Nota de Crédito"
    hrc_template = pd.concat([hrc_template, nota_credito], ignore_index=True)

    columnas_finales = [
        "Tipo de Documento",
        "Referencia / Factura",
        "Importe de factura",
        "Descuento",
        "Motivo del descuento",
        "Pago Neto",
        "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # --- Paso 5: Exportar con formato y colores directo ---
    hrc_template = hrc_template.sort_values(by="Tipo de Documento", ascending=False).reset_index(drop=True)
    hrc_template.to_excel(rutas["salida"], index=False, sheet_name="Template", startrow=17, startcol=2)

    # Abrir para aplicar estilos
    wb = load_workbook(rutas["salida"])
    ws = wb["Template"]

    azul_oscuro = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    celeste_intenso = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    letra_blanca = Font(color="FFFFFF")
    letra_negra = Font(color="000000")

    # Encabezados
    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = ""
    ws["C6"] = "Cliente"
    ws["D6"] = ""
    ws["C8"] = "Codigo de Cliente"
    ws["D8"] = ""
    ws["C11"] = "Referencia"
    ws["C12"] = ""
    ws["D11"] = "Fecha"
    ws["D12"] = ""
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = 0.0
    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = 0.0
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = hrc_template["Pago Neto"].sum()
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = hrc_template["Pago Neto"].sum() - 0.0

    # Colores
    for cell in ["G2", "C6", "C7", "C8", "F6", "F7", "F8", "C11", "D11", "E11",
                 "F11", "C18", "D18", "E18", "F18", "G18", "H18", "I18"]:
        if ws[cell].value is not None:
            ws[cell].fill = azul_oscuro
            ws[cell].font = letra_blanca

    for cell in ["C4", "D6", "D7", "D8", "G6", "G7", "G8"]:
        if ws[cell].value is not None:
            ws[cell].fill = celeste_intenso
            ws[cell].font = letra_blanca

    ws["H2"].fill = amarillo
    ws["H2"].font = letra_negra

    # Formatos numéricos
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'

    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(rutas["salida"])
    print(f"✅ Archivo exportado con formato y colores: {rutas['salida']}")

    return hrc_template
