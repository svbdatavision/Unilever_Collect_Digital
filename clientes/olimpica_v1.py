# clientes/olimpica.py

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def procesar():
    # --- LEER EXCELS ---
    remittance = pd.read_excel(
        "Remittance.xlsx",
        skiprows=25,
        nrows=65,
        usecols=["DOC", "No. Doc", "Total a Pagar"]
    ).dropna(subset=["DOC"])

    FBL5N = pd.read_excel(
        "FBL5N.xlsx",
        usecols=["Document Type", "Reference", "Amount in local currency"]
    )

    wb_fbl3n = load_workbook("FBL3N.xlsx", data_only=True)
    ws_fbl3n = wb_fbl3n.active
    fecha_original = ws_fbl3n["K8"].value
    fecha_dt = datetime.strptime(fecha_original, "%d.%m.%Y")
    fecha_pago = fecha_dt.strftime("%-m/%-d/%y")  # Windows: "%#m/%#d/%y"
    importe_FBL3N = ws_fbl3n["N8"].value
    importe_FBL3N = abs(float(str(importe_FBL3N).replace(".", "").replace(",", ".")))

    # --- TRANSFORMACIONES REMITTANCE ---
    remittance = remittance.rename(columns={
        "DOC": "Tipo de Documento",
        "No. Doc": "Referencia / Factura",
        "Total a Pagar": "Importe de factura"
    })
    remittance["Referencia / Factura"] = remittance["Referencia / Factura"].str[1:-3]
    remittance["Tipo de Documento"] = remittance["Tipo de Documento"].replace(
        {"Factura Comercial": "Factura", "Nota Crédito": "Descuentos no asociados a FC"}
    )
    remittance["Importe de factura"] = pd.to_numeric(remittance["Importe de factura"], errors="coerce").round(2)

    # Columnas vacías si no existen
    for col in ["Descuento", "Motivo del descuento", "Pago Neto", "Comentarios"]:
        if col not in remittance.columns:
            remittance[col] = ""

    # Reglas de descuentos
    conds = [
        (remittance["Referencia / Factura"].str.contains("XKZV", na=False)),
        (remittance["Referencia / Factura"].str.startswith("PMP", na=False) & (remittance["Importe de factura"] < 0))
        # agregá otras condiciones según necesites
    ]
    descuentos = ["FACT PROVEEDOR", "RECHAZO"]
    motivos = ["CSB", "551"]

    remittance["Descuento"] = np.select(conds, descuentos, default=remittance["Descuento"])
    remittance["Motivo del descuento"] = np.select(conds, motivos, default=remittance["Motivo del descuento"])

    # --- TRANSFORMACIONES FBL5N ---
    FBL5N = FBL5N[FBL5N["Document Type"].isin(["RV", "NRO"])]
    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    # --- MERGE ---
    hrc_template = pd.merge(remittance, FBL5N, on="Referencia / Factura", how="left")

    # --- CÁLCULOS ---
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )

    diferencias = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()
    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": diferencias["Referencia / Factura"],
        "Importe de factura": diferencias["Diferencia"],
        "Pago Neto": diferencias["Diferencia"],
        "Descuento": diferencias["Diferencia"].apply(lambda x: "MENORES VALORES" if -20000 < x < 20000 else "A definir"),
        "Motivo del descuento": diferencias["Diferencia"].apply(lambda x: "WOB" if x < 0 else "384"),
        "Comentarios": ""
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # --- COLUMNAS DINÁMICAS ---
    hrc_template["Comentarios"] = hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("")
    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    columnas_finales = [
        "Tipo de Documento", "Referencia / Factura", "Importe de factura",
        "Descuento", "Motivo del descuento", "Pago Neto", "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # --- EXPORTACIÓN ---
    ruta_salida = "Template_HRC_olimpica.xlsx"
    hrc_template.to_excel(ruta_salida, index=False, sheet_name="Template", startrow=17, startcol=2)
    wb = load_workbook(ruta_salida)
    ws = wb["Template"]

    # Títulos
    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = ""  # Si tenés numero_orden, insertalo
    ws["C6"] = "Cliente"
    ws["C8"] = "Codigo de Cliente"
    ws["D6"] = ""  # nombre_cliente
    ws["D8"] = ""  # id_cliente
    ws["C11"] = "Referencia"
    ws["C12"] = ""  # numero_orden
    ws["D11"] = "Fecha"
    ws["D12"] = fecha_pago
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = importe_FBL3N
    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = importe_FBL3N
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = hrc_template["Pago Neto"].sum()
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = -(hrc_template["Pago Neto"].sum() - importe_FBL3N)

    # Formatos
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'

    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(ruta_salida)
    print(f"Archivo exportado correctamente: {ruta_salida}")

    return hrc_template
