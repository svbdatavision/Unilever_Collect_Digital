import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def procesar():
    # --- Paso 1: Leer Remittance.xlsx ---
    remittance = pd.read_excel(
        "Remittance.xlsx",
        skiprows=25,
        nrows=65,
        usecols=["DOC", "No. Doc", "Total a Pagar"]
    ).dropna(subset=["DOC"])

    # Transformaciones Remittance
    remittance = remittance.rename(columns={
        "DOC": "Tipo de Documento",
        "No. Doc": "Referencia / Factura",
        "Total a Pagar": "Importe de factura"
    })

    remittance["Referencia / Factura"] = remittance["Referencia / Factura"].str[1:-3]
    remittance["Tipo de Documento"] = remittance["Tipo de Documento"].replace({
        "Factura Comercial": "Factura",
        "Nota Crédito": "Descuentos no asociados a FC"
    })
    remittance["Importe de factura"] = pd.to_numeric(remittance["Importe de factura"], errors="coerce").round(2)

    # Columnas vacías
    for col in ["Descuento", "Motivo del descuento"]:
        if col not in remittance.columns:
            remittance[col] = ""

    # Reglas de descuentos
    conds = [
        (remittance["Referencia / Factura"].str.startswith("PMP", na=False)) & (remittance["Importe de factura"] < 0),
        remittance["Referencia / Factura"].str.startswith("0085489", na=False),
        remittance["Referencia / Factura"].str.startswith("463", na=False),
        remittance["Referencia / Factura"].str.startswith("9801649", na=False),
        remittance["Referencia / Factura"].str.startswith("310", na=False)
    ]
    descuentos = ["RECHAZO", "DESCUENTO", "AVERIA", "FACT PROVEEDOR", "NOTA DEBITO"]
    motivos = ["551", "987", "522", "CSB", "987"]
    remittance["Descuento"] = np.select(conds, descuentos, default=remittance["Descuento"])
    remittance["Motivo del descuento"] = np.select(conds, motivos, default=remittance["Motivo del descuento"])

    # --- Paso 2: Leer FBL5N.xlsx ---
    FBL5N = pd.read_excel(
        "FBL5N.xlsx",
        usecols=["Document Type", "Reference", "Amount in local currency"]
    )
    FBL5N = FBL5N[FBL5N["Document Type"] == "RV"]
    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    # --- Paso 3: Merge ---
    hrc_template = pd.merge(
        remittance,
        FBL5N,
        on="Referencia / Factura",
        how="left"
    )

    # Diferencias
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )

    diferencias = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()
    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": diferencias["Referencia / Factura"],
        "Importe de factura": diferencias["Diferencia"],
        "Pago Neto": "",
        "Descuento": diferencias["Diferencia"].apply(lambda x: "MENORES VALORES" if -20000 < x < 20000 else "A definir"),
        "Motivo del descuento": diferencias["Diferencia"].apply(lambda x: "WOB" if x < 0 else "384")
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # Comentarios y Pago Neto
    hrc_template["Comentarios"] = hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("")
    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    # Columnas finales
    columnas_finales = [
        "Tipo de Documento",
        "Referencia / Factura",
        "Importe de factura",
        "Descuento",
        "Motivo del descuento",
        "Pago Neto",
        "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # --- Paso 4: Datos dinámicos ---
    # Numero de orden
    wb_rem = load_workbook("Remittance.xlsx", data_only=True)
    ws_rem = wb_rem.active
    celda = ws_rem["C10"].value
    numero_orden = celda.split("Orden de Pago:")[1].strip() if celda and "Orden de Pago:" in celda else ""

    # Cliente
    fbl5n = pd.read_excel("FBL5N.xlsx", usecols=["Customer", "Name 1"], nrows=1)
    id_cliente = fbl5n["Customer"].iloc[0]
    nombre_cliente = fbl5n["Name 1"].iloc[0]

    # Fecha y valor de pago (FBL3N)
    wb_fbl3n = load_workbook("FBL3N.xlsx", data_only=True)
    ws_fbl3n = wb_fbl3n.active
    fecha_dt = datetime.strptime(ws_fbl3n["K8"].value, "%d.%m.%Y")
    fecha_pago = fecha_dt.strftime("%-m/%-d/%y")  # Linux/macOS
    importe_FBL3N = abs(float(ws_fbl3n["N8"].value.replace(".", "").replace(",", ".")))

    # --- Paso 5: Sumas ---
    total_pago_neto = hrc_template["Pago Neto"].sum()
    diferencia = total_pago_neto - importe_FBL3N

    # --- Paso 6: Exportar Excel ---
    ruta_salida = "Template_HRC_olimpica.xlsx"
    hrc_template.to_excel(ruta_salida, index=False, sheet_name="Template", startrow=17, startcol=2)

    # --- Paso 7: Formato con openpyxl ---
    wb = load_workbook(ruta_salida)
    ws = wb["Template"]

    # Titulos y cuadros
    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = numero_orden  # dato dinámico
    ws["C6"] = "Cliente"
    ws["C8"] = "Codigo de Cliente"
    ws["D6"] = nombre_cliente  # dato dinámico
    ws["D8"] = id_cliente       # dato dinámico
    ws["C11"] = "Referencia"
    ws["C12"] = numero_orden    # dato dinámico
    ws["D11"] = "Fecha"
    ws["D12"] = fecha_pago      # dato dinámico
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = importe_FBL3N

    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = importe_FBL3N
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = total_pago_neto
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = -diferencia

    # Formato numérico
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'

    # Formato tabla
    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(ruta_salida)
    print(f"Archivo exportado correctamente: {ruta_salida}")

    return hrc_template
