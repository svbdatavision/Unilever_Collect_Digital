import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from copy import copy
import os

def procesar():
    # --- Definir rutas como objetos ---
    rutas = {
        "remittance": os.path.join("Archivos", "Remittance", "Remittance_olimpica.xlsx"),
        "fbl5n": os.path.join("Archivos", "Base_de_datos", "FBL5N_olimpica.xlsx"),
        "fbl3n": os.path.join("Archivos", "Base_de_datos", "FBL3N.xlsx"),
        "salida": os.path.join("Archivos", "Template", "Template_HRC_olimpica.xlsx")
    }

    # --- Paso 1: Leer Remittance.xlsx ---
    remittance = pd.read_excel(
        rutas["remittance"],
        skiprows=25,
        nrows=65,
        usecols=["DOC", "No. Doc", "Total a Pagar"]
    ).dropna(subset=["DOC"])

    # Transformaciones Remittance
    remittance = remittance.rename(columns={
        "DOC": "Tipo de Documento",
        "No. Doc": "Referencia / Factura",
        "Total a Pagar": "Importe de factura"
    })
    remittance["Referencia / Factura"] = remittance["Referencia / Factura"].str[1:-3]
    remittance["Tipo de Documento"] = remittance["Tipo de Documento"].replace({
        "Factura Comercial": "Factura",
        "Nota Crédito": "Descuentos no asociados a FC"
    })
    remittance["Importe de factura"] = pd.to_numeric(remittance["Importe de factura"], errors="coerce").round(2)

    # Columnas vacías
    for col in ["Descuento", "Motivo del descuento"]:
        if col not in remittance.columns:
            remittance[col] = ""

    # Reglas de descuentos
    conds = [
        (remittance["Referencia / Factura"].str.startswith("PMP", na=False)) & (remittance["Importe de factura"] < 0),
        remittance["Referencia / Factura"].str.startswith("0085489", na=False),
        remittance["Referencia / Factura"].str.startswith("463", na=False),
        remittance["Referencia / Factura"].str.startswith("9801649", na=False),
        remittance["Referencia / Factura"].str.startswith("310", na=False)
    ]
    descuentos = ["RECHAZO", "DESCUENTO", "AVERIA", "FACT PROVEEDOR", "NOTA DEBITO"]
    motivos = ["551", "987", "522", "CSB", "987"]
    remittance["Descuento"] = np.select(conds, descuentos, default=remittance["Descuento"])
    remittance["Motivo del descuento"] = np.select(conds, motivos, default=remittance["Motivo del descuento"])

    # --- Paso 2: Leer FBL5N.xlsx ---
    FBL5N = pd.read_excel(
        rutas["fbl5n"],
        usecols=["Document Type", "Reference", "Amount in local currency"]
    )
    FBL5N = FBL5N[FBL5N["Document Type"] == "RV"]
    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    # --- Paso 3: Merge ---
    hrc_template = pd.merge(remittance, FBL5N, on="Referencia / Factura", how="left")

    # Diferencias
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )
    diferencias = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()
    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": diferencias["Referencia / Factura"],
        "Importe de factura": diferencias["Diferencia"],
        "Pago Neto": "",
        "Descuento": diferencias["Diferencia"].apply(lambda x: "MENORES VALORES" if -20000 < x < 20000 else "A definir"),
        "Motivo del descuento": diferencias["Diferencia"].apply(lambda x: "WOB" if x < 0 else "384")
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # --- Comentarios y Pago Neto ---
    hrc_template["Comentarios"] = np.where(
        hrc_template["Tipo de Documento"] == "Factura",
        "",
        np.where(
            hrc_template["Descuento"] == "MENORES VALORES",
            hrc_template["Descuento"],
            hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("")
        )
    )
    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    columnas_finales = [
        "Tipo de Documento", "Referencia / Factura", "Importe de factura",
        "Descuento", "Motivo del descuento", "Pago Neto", "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # --- Paso 4: Datos dinámicos ---
    wb_rem = load_workbook(rutas["remittance"], data_only=True)
    ws_rem = wb_rem.active
    celda = ws_rem["C10"].value
    numero_orden = celda.split("Orden de Pago:")[1].strip() if celda and "Orden de Pago:" in celda else ""

    fbl5n = pd.read_excel(rutas["fbl5n"], usecols=["Customer", "Name 1"], nrows=1)
    id_cliente = fbl5n["Customer"].iloc[0]
    nombre_cliente = fbl5n["Name 1"].iloc[0]

    wb_fbl3n = load_workbook(rutas["fbl3n"], data_only=True)
    ws_fbl3n = wb_fbl3n.active
    fecha_dt = datetime.strptime(ws_fbl3n["K8"].value, "%d.%m.%Y")
    fecha_pago = fecha_dt.strftime("%-m/%-d/%y")
    importe_FBL3N = abs(float(ws_fbl3n["N8"].value.replace(".", "").replace(",", ".")))

    total_pago_neto = hrc_template["Pago Neto"].sum()
    diferencia = total_pago_neto - importe_FBL3N

    # --- Paso 5: Exportar Excel ---
    with pd.ExcelWriter(rutas["salida"], engine="openpyxl") as writer:
        hrc_template.to_excel(writer, index=False, sheet_name="Template", startrow=17, startcol=2)

    wb_salida = load_workbook(rutas["salida"])
    ws = wb_salida["Template"]

    # --- Aplicar formatos hoja Template ---
    azul_oscuro = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    celeste_intenso = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    letra_blanca = Font(color="FFFFFF")
    letra_negra = Font(color="000000")

    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = numero_orden
    ws["C6"] = "Cliente"
    ws["C8"] = "Codigo de Cliente"
    ws["D6"] = nombre_cliente
    ws["D8"] = id_cliente
    ws["C11"] = "Referencia"
    ws["C12"] = numero_orden
    ws["D11"] = "Fecha"
    ws["D12"] = fecha_pago
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = importe_FBL3N

    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = importe_FBL3N
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = total_pago_neto
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = -diferencia

    # Formato numérico
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'

    # Formato tabla
    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ["G2", "C6", "C7", "C8", "F6", "F7", "F8", "C11", "D11", "E11",
                 "F11", "C18", "D18", "E18", "F18", "G18", "H18", "I18"]:
        ws[cell].fill = azul_oscuro
        ws[cell].font = letra_blanca
    for cell in ["C4", "D6", "D7", "D8", "G6", "G7", "G8"]:
        ws[cell].fill = celeste_intenso
        ws[cell].font = letra_blanca
    ws["H2"].fill = amarillo
    ws["H2"].font = letra_negra

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # --- Copiar hoja Remittance exacta ---
    wb_remittance = load_workbook(rutas["remittance"], data_only=False)
    ws_original = wb_remittance.active
    ws_nueva = wb_salida.create_sheet("Remittance")

    for row in ws_original.iter_rows():
        for cell in row:
            new_cell = ws_nueva.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)

    for col_letter, col_dim in ws_original.column_dimensions.items():
        ws_nueva.column_dimensions[col_letter].width = col_dim.width
    for row_idx, row_dim in ws_original.row_dimensions.items():
        ws_nueva.row_dimensions[row_idx].height = row_dim.height

    wb_salida.save(rutas["salida"])
    print(f"Archivo exportado correctamente con formato: {rutas['salida']}")

    return hrc_template
