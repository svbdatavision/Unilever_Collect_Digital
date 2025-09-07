import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Font

def procesar():
    import os

    # --- Definir rutas ---
    ruta_remittance = os.path.join("Archivos", "Remittance", "Remittance_farmatodo.xlsx")
    ruta_fbl5n = os.path.join("Archivos", "Base_de_datos", "FBL5N_farmatodo.xlsx")
    ruta_salida = os.path.join("Archivos", "Template", "Template_HRC_farmatodo.xlsx")

    # --- Cargar Remittance ---
    remittance = pd.read_excel(
        ruta_remittance,
        skiprows=6,
        nrows=20,
        header=[0, 1]
    )

    # Aplanar columnas multi-index
    remittance.columns = [
        str(col[0]).strip() if "Unnamed" in str(col[1]) else str(col[1]).strip()
        for col in remittance.columns
    ]
    remittance = remittance[["Nro Factura", "Descripción", "Total"]]
    remittance = remittance.rename(columns={
        "Nro Factura": "Referencia / Factura",
        "Total": "Importe de factura"
    })
    remittance["Importe de factura"] = pd.to_numeric(remittance["Importe de factura"], errors="coerce").round(2)

    # --- Tipo de Documento ---
    conds = [
        remittance["Referencia / Factura"].str.startswith("PMP", na=False) & (remittance["Importe de factura"] > 0),
        ~remittance["Referencia / Factura"].str.startswith("PMP", na=False) & (remittance["Importe de factura"] > 0),
        remittance["Importe de factura"] < 0
    ]
    choices = ["Factura", "Nota Debito", "Descuentos no asociados a FC"]
    remittance["Tipo de Documento"] = np.select(conds, choices, default="")

    # --- Descuento y Motivo ---
    if "Descuento" not in remittance.columns:
        remittance["Descuento"] = ""
    if "Motivo del descuento" not in remittance.columns:
        remittance["Motivo del descuento"] = ""

    conds_desc = [
        remittance["Referencia / Factura"].str.startswith("NC-DC05", na=False),
        remittance["Referencia / Factura"].str.startswith("NC-DC04", na=False),
        remittance["Referencia / Factura"].str.startswith("NC-DC06", na=False),
        remittance["Referencia / Factura"].str.contains("XKZV", na=False),
        remittance["Referencia / Factura"].str.startswith("NCF-DC", na=False),
        remittance["Referencia / Factura"].str.startswith("NC-PMP", na=False),
        remittance["Referencia / Factura"].str.startswith("CNQPMP", na=False),
        remittance["Referencia / Factura"].str.startswith("NC-100", na=False),
        remittance["Referencia / Factura"].str.startswith("NC1", na=False)
    ]
    descuentos = ["CONVENIOS", "DSCT PROMOCIONAL", "DSCT PROMOCIONAL", "FACT PROVEEDOR",
                  "CONVENIOS", "DSCT AVERIAS", "PGO PDTE SOPORTE", "DSCT AVERIAS", "DSCT AVERIAS"]
    motivos = ["657", "987", "987", "CSB", "657", "206", "551", "522", "522"]

    remittance["Descuento"] = np.select(conds_desc, descuentos, default=remittance["Descuento"])
    remittance["Motivo del descuento"] = np.select(conds_desc, motivos, default=remittance["Motivo del descuento"])

    # --- Ordenar Tipo de Documento ---
    remittance = remittance.sort_values(by="Tipo de Documento", ascending=False).reset_index(drop=True)

    # --- Cargar FBL5N ---
    FBL5N = pd.read_excel(
        ruta_fbl5n,
        sheet_name="Sheet1",
        usecols=["Document Type", "Reference", "Amount in local currency", "Reason code", "Document Number", "Text"]
    )
    FBL5N = FBL5N[(FBL5N["Document Type"] == "RV") | (FBL5N["Reason code"] == "NRO")]
    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    # --- Ajustar NRO ---
    FBL5N["Referencia / Factura"] = np.where(
        FBL5N["Reason code"] == "NRO",
        FBL5N["Document Number"].astype("Int64").astype(str),
        FBL5N["Referencia / Factura"]
    )

    # --- Merge ---
    hrc_template = pd.merge(
        remittance,
        FBL5N,
        on="Referencia / Factura",
        how="left"
    )

    # --- Diferencias ---
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )

    diferencias = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()
    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": diferencias["Referencia / Factura"],
        "Importe de factura": diferencias["Diferencia"],
        "Pago Neto": "",
        "Descuento": diferencias["Diferencia"].apply(lambda x: "MENORES VALORES" if -20000 < x < 20000 else "A definir"),
        "Motivo del descuento": diferencias["Diferencia"].apply(lambda x: "WOB" if x < 0 else "384")
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # --- Comentarios y Pago Neto ---
    hrc_template["Comentarios"] = np.where(
        hrc_template["Tipo de Documento"] == "Factura",
        "",
        np.where(
            hrc_template["Descuento"] == "MENORES VALORES",
            hrc_template["Descuento"],
            hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("") + " " + hrc_template["Descripción"].fillna("")
        )
    )
    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    # --- Nota de Crédito ---
    nota_credito = FBL5N[FBL5N["Reason code"] == "NRO"].copy()
    nota_credito["Tipo de Documento"] = "Nota de Crédito"
    hrc_template = pd.concat([hrc_template, nota_credito], ignore_index=True)
    hrc_template.loc[
        hrc_template["Tipo de Documento"] == "Nota de Crédito",
        ["Importe de factura", "Pago Neto", "Comentarios", "Motivo del descuento"]
    ] = hrc_template.loc[
        hrc_template["Tipo de Documento"] == "Nota de Crédito",
        ["importe_FBL5N", "importe_FBL5N", "Text", "Reason code"]
    ].values

    # --- Columnas finales ---
    columnas_finales = [
        "Tipo de Documento", "Referencia / Factura", "Importe de factura",
        "Descuento", "Motivo del descuento", "Pago Neto", "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # --- Datos dinámicos ---
    wb_rem = load_workbook(ruta_remittance, data_only=True)
    ws_rem = wb_rem.active  
    numero_orden = ws_rem["B7"].value

    fbl5n = pd.read_excel(ruta_fbl5n, usecols=["Customer", "Name 1"], nrows=1)
    id_cliente = fbl5n["Customer"].iloc[0]
    nombre_cliente = fbl5n["Name 1"].iloc[0]

    fecha_pago = datetime.today().date()
    importe_FBL3N = 1  # placeholder
    total_pago_neto = hrc_template["Pago Neto"].sum()
    diferencia = total_pago_neto - importe_FBL3N

    # --- Exportar Excel ---
    hrc_template.to_excel(ruta_salida, index=False, sheet_name="Template", startrow=17, startcol=2)

    # --- Abrir para aplicar formatos ---
    wb = load_workbook(ruta_salida)
    ws = wb["Template"]

    # Estilos
    azul_oscuro = PatternFill(start_color="002366", end_color="002366", fill_type="solid")
    celeste_intenso = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    letra_blanca = Font(color="FFFFFF")
    letra_negra = Font(color="000000")

    # --- Asignar valores a celdas ---
    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = numero_orden
    ws["C6"] = "Cliente"
    ws["C8"] = "Codigo de Cliente"
    ws["D6"] = nombre_cliente
    ws["D8"] = id_cliente
    ws["C11"] = "Referencia"
    ws["C12"] = numero_orden
    ws["D11"] = "Fecha"
    ws["D12"] = fecha_pago
    ws["E11"] = "Método de Pago"
    ws["E12"] = "Transferencia"
    ws["F11"] = "Valor"
    ws["F12"] = importe_FBL3N
    ws["F6"] = "TOTAL s/ BANCOS"
    ws["G6"] = importe_FBL3N
    ws["F7"] = "TOTAL s/ DETALLE"
    ws["G7"] = total_pago_neto
    ws["F8"] = "DIFERENCIA"
    ws["G8"] = -diferencia

    # --- Formato celdas ---
    for cell in ["G2","C6","C7","C8","F6","F7","F8","C11","D11","E11",
                 "F11","C18","D18","E18","F18","G18","H18","I18"]:
        ws[cell].fill = azul_oscuro
        ws[cell].font = letra_blanca
    for cell in ["C4","D6","D7","D8","G6","G7","G8"]:
        ws[cell].fill = celeste_intenso
        ws[cell].font = letra_blanca
    ws["H2"].fill = amarillo
    ws["H2"].font = letra_negra

    # Formato tabla principal
    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Ajustar ancho de columnas automáticamente ---
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
                
    wb.save(ruta_salida)
    print(f"Archivo exportado correctamente con formato: {ruta_salida}")

    return hrc_template
