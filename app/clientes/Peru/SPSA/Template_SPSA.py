import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font,  Border, Side



# Rutas
ruta_entrada = fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\SPSA\Remmitance_SPSA.xlsx"
ruta_salida = fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\SPSA\formato_resultado.xlsx"

# Cargar archivo de remesas
remittance_df = pd.read_excel(ruta_entrada)
remittance_df.columns = remittance_df.columns.str.strip()

# Mapeo de tipos
tipo_map = {
    "Nota de Débito": "ND",
    "Nota de Crédito": "NC",
    "Factura Mercadería": "Factura",
    "Facturas por Cobrar": "Fact. Convenio",
}

# Transformación
formato_df = pd.DataFrame()
formato_df["Tipo Doc"] = remittance_df["Descripción Tipo Documento"].map(lambda x: tipo_map.get(x, "Otro"))
formato_df["Referencia / Factura"] = remittance_df["Nro. Documento Proveedor"]
formato_df["Importe de factura"] = remittance_df["Importe Documento"]

def razon_descuento(row):
    ref = str(row["Nro. Documento Proveedor"])
    if ref.startswith("F0") or ref.startswith("FN"):
        return "657"
    if ref.startswith("F701") or ref.startswith("F121"):
        return ""
    return tipo_map.get(row["Tipo"], "Otro")

formato_df["Razon de Descuento"] = remittance_df.apply(razon_descuento, axis=1)
formato_df["Importe de factura"] = formato_df["Importe de factura"].round(2)
formato_df = formato_df.dropna(subset=["Referencia / Factura", "Importe de factura"])

# Exportar a Excel
formato_df.to_excel(ruta_salida, index=False)

# Insertar datos en columnas F y G
wb = load_workbook(ruta_salida)
ws = wb.active

# Datos a insertar
datos_extra = [
    ("Nombre Cliente", "SUPERM PERUANOS S A"),
    ("Numero de Cliente", "10263165"),
    ("Referencia de Pago", ""),
    ("Pago", sum(formato_df["Importe de factura"])),
    ("Metodo de Pago", "TRANSFERENCIA"),
    ("Fecha de pago", ""),
]

# Estilo: fondo azul claro y texto en negrita
fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
font_negrita = Font(bold=True)

# Bordes negros
borde_negro = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# Insertar y aplicar estilo
for i, (col_f, col_g) in enumerate(datos_extra, start=1):
    celda_f = ws[f"F{i}"]
    celda_g = ws[f"G{i}"]

    celda_f.value = col_f
    celda_g.value = col_g

    celda_f.fill = fill_azul
    celda_f.font = font_negrita
    celda_f.border = borde_negro
    celda_g.border = borde_negro


wb.save(ruta_salida)
print("✅ Archivo actualizado con datos en columnas F y G")

