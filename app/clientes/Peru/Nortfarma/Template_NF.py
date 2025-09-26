import pdfplumber
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font,  Border, Side

user = os.getenv("USERNAME")

# Ruta al archivo PDF
pdf_path = fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\Nortfarma\Remmitance nortfarma.pdf"
output_path = fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\Nortfarma\resultado_convertido.xlsx"

# Extraer las filas de la tabla
rows = []
with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                rows.append(row)

# Buscar encabezado
header_idx = None
for i, row in enumerate(rows):
    if row and "Fec Emision" in row:
        header_idx = i
        break

# Procesar datos
header = rows[header_idx]
data_rows = rows[header_idx + 1:]
df = pd.DataFrame(data_rows, columns=header)

# Eliminar columna 'Fec Emision'
df = df.drop(columns=["Fec Emision"])

# Renombrar columnas
df = df.rename(columns={
    "Num Documento": "Referencia / Factura",
    "Monto": "Importe"
})

# Agregar columna 'Razon de Descuento'
df["Razon de Descuento"] = ""
df.loc[df["Tipo Doc"].str.upper() == "FACTURA X COBRAR", "Razon de Descuento"] = "657"

# Formatear 'Referencia / Factura' con ceros a la izquierda si es FACTURA o NOTA DE CREDITO
def pad_correlativo(ref, tipo):
    if not isinstance(ref, str):
        return ref
    if tipo.upper() not in ["FACTURA", "NOTA DE CREDITO"]:
        return ref
    m = re.match(r"^([A-Z0-9]+)-(\d+)$", ref)
    if m:
        prefijo, correlativo = m.groups()
        correlativo_padded = correlativo.zfill(8)
        return f"{prefijo}-{correlativo_padded}"
    return ref

df["Referencia / Factura"] = [
    pad_correlativo(ref, tipo)
    for ref, tipo in zip(df["Referencia / Factura"], df["Tipo Doc"])
]

# Convertir 'Importe' a número y aplicar signo según tipo de documento
df["Importe"] = df["Importe"].str.replace(",", "")
df["Importe"] = pd.to_numeric(df["Importe"], errors="coerce")

df.loc[df["Tipo Doc"].str.upper() == "FACTURA", "Importe"] *= 1
df.loc[df["Tipo Doc"].str.upper().isin(["NOTA DE CREDITO", "FACTURA X COBRAR"]), "Importe"] *= -1

# Reordenar columnas
df_final = df[["Tipo Doc", "Referencia / Factura", "Importe", "Razon de Descuento"]]

# Datos a insertar
datos_extra = [
    ("Nombre Cliente", "NORTFARMA S A C"),
    ("Numero de Cliente", "10449763"),
    ("Referencia de Pago", ""),
    ("Pago", sum(df_final["Importe"])),
    ("Metodo de Pago", "TRANSFERENCIA"),
    ("Fecha de pago", ""),
]

# Guardar en Excel
df_final.to_excel(output_path, index=False)

wb = load_workbook(output_path)
ws = wb.active

# Estilo: fondo azul claro y texto en negrita
fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
font_negrita = Font(bold=True)

# Bordes negros
borde_negro = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# Insertar y aplicar estilo
for i, (col_f, col_g) in enumerate(datos_extra, start=1):
    celda_f = ws[f"F{i}"]
    celda_g = ws[f"G{i}"]

    celda_f.value = col_f
    celda_g.value = col_g

    celda_f.fill = fill_azul
    celda_f.font = font_negrita
    celda_f.border = borde_negro
    celda_g.border = borde_negro

wb.save(output_path)