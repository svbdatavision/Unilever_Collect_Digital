import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font,  Border, Side

user = os.getenv("USERNAME")

# Cargar el archivo de remesas
remittance_df = pd.read_excel(fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\Cencosud\Remmitance_Cencosud.xlsx")

# Limpiar espacios en los nombres de las columnas
remittance_df.columns = remittance_df.columns.str.strip()

# Reemplazar los valores en la columna correcta (ajusta el nombre si es necesario)
if "Nro. Documento" in remittance_df.columns:
    remittance_df["Nro. Documento"] = remittance_df["Nro. Documento"].str.replace(r'^(01|07|08|00)-F', 'F', regex=True)

# Mapeo de tipos de documento a razones de descuento
tipo_map = {
    "NOTA DE CRÉDITO PROVEEDOR": "NC",
    "NOTA DE CREDITO PROVEEDOR": "NC",
    "FACTURA PROVEEDOR": "Factura",
    "Auto-Detracción Deud": "Fact. Convenio",
    "Auto-Detracci?n Deud": "Fact. Convenio",
    "Pago detracción": "Fact. Convenio",
    "FACTURA DEUDOR METRO": "Fact. Convenio"
}

# Aplicar transformación
formato_df = pd.DataFrame()
formato_df["Tipo Doc"] = remittance_df["Tipo"].map(lambda x: tipo_map.get(x, "Otro"))
formato_df["Referencia / Factura"] = remittance_df["Nro. Documento"]
formato_df["Importe de factura"] = remittance_df["Importe"]

# Crear la columna "Razon de Descuento" con el mapeo y la condición adicional
def razon_descuento(row):
    ref = str(row["Nro. Documento"])
    if ref.startswith("FA") or ref.startswith("FN"):
        return "657"
    if ref.startswith("F701") or ref.startswith("F121"):
        return ""
    return tipo_map.get(row["Tipo"], "Otro")

formato_df["Razon de Descuento"] = remittance_df.apply(razon_descuento, axis=1)
ruta_salida = fr"C:\Users\{user}\Unilever\Codigos - Documents\Automatizaciones MT\Template\Cencosud\formato_resultado.xlsx"

# Si existe el archivo, lo elimina antes de guardar el nuevo
if os.path.exists(ruta_salida):
    os.remove(ruta_salida)

# Exportar el DataFrame resultante a un archivo Excel
formato_df.to_excel(ruta_salida, index=False)
# Insertar datos en columnas F y G
wb = load_workbook(ruta_salida)
ws = wb.active

# Datos a insertar
datos_extra = [
    ("Nombre Cliente", "CENCOSUD RETAIL PERU S A"),
    ("Numero de Cliente", "10262842"),
    ("Referencia de Pago", ""),
    ("Pago", sum(formato_df["Importe de factura"])),
    ("Metodo de Pago", "TRANSFERENCIA"),
    ("Fecha de pago", ""),
]

# Estilo: fondo azul claro y texto en negrita
fill_azul = PatternFill(start_color="D0E9F8", end_color="D0E9F8", fill_type="solid")
font_negrita = Font(bold=True)

# Bordes negros
borde_negro = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# Insertar y aplicar estilo
for i, (col_f, col_g) in enumerate(datos_extra, start=1):
    celda_f = ws[f"F{i}"]
    celda_g = ws[f"G{i}"]

    celda_f.value = col_f
    celda_g.value = col_g

    celda_f.fill = fill_azul
    celda_f.font = font_negrita
    celda_f.border = borde_negro
    celda_g.border = borde_negro


wb.save(ruta_salida)

print("✅ Archivo exportado como formato_resultado.xlsx")
