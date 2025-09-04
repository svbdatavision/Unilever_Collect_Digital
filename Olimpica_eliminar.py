# olimpica.py
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def procesar_olimpica(remittance_file="Remittance.xlsx", fbl5n_file="FBL5N.xlsx", fbl3n_file="FBL3N.xlsx"):
    # -------------------
    # Paso 1: Leer remittance
    # -------------------
    remittance = pd.read_excel(
        remittance_file,
        skiprows=25,
        nrows=65,
        usecols=["DOC", "No. Doc", "Total a Pagar"]
    ).dropna(subset=["DOC"])

    # Transformaciones iniciales
    remittance = remittance.rename(columns={
        "DOC": "Tipo de Documento",
        "No. Doc": "Referencia / Factura",
        "Total a Pagar": "Importe de factura"
    })

    remittance["Referencia / Factura"] = remittance["Referencia / Factura"].str[1:-3]
    remittance["Tipo de Documento"] = remittance["Tipo de Documento"].replace({
        "Factura Comercial": "Factura",
        "Nota Crédito": "Descuentos no asociados a FC"
    })
    remittance["Importe de factura"] = pd.to_numeric(remittance["Importe de factura"], errors="coerce").round(2)

    # -------------------
    # Paso 2: Crear columnas Descuento y Motivo
    # -------------------
    if "Descuento" not in remittance.columns:
        remittance["Descuento"] = ""
    if "Motivo del descuento" not in remittance.columns:
        remittance["Motivo del descuento"] = ""

    conds = [
        (remittance["Referencia / Factura"].str.startswith("PMP", na=False)) & (remittance["Importe de factura"] < 0),
        remittance["Referencia / Factura"].str.startswith("0085489", na=False),
        remittance["Referencia / Factura"].str.startswith("463", na=False),
        remittance["Referencia / Factura"].str.startswith("9801649", na=False),
        remittance["Referencia / Factura"].str.startswith("310", na=False)
    ]
    descuentos = ["RECHAZO", "DESCUENTO", "AVERIA", "FACT PROVEEDOR", "NOTA DEBITO"]
    motivos = ["551", "987", "522", "CSB", "987"]

    remittance["Descuento"] = np.select(conds, descuentos, default=remittance["Descuento"])
    remittance["Motivo del descuento"] = np.select(conds, motivos, default=remittance["Motivo del descuento"])

    # -------------------
    # Paso 3: Leer FBL5N
    # -------------------
    FBL5N = pd.read_excel(
        fbl5n_file,
        usecols=["Document Type", "Reference", "Amount in local currency"]
    )
    FBL5N = FBL5N[FBL5N["Document Type"].isin(["RV", "NRO"])]
    FBL5N = FBL5N.rename(columns={
        "Reference": "Referencia / Factura",
        "Amount in local currency": "importe_FBL5N"
    }).reset_index(drop=True)

    # -------------------
    # Paso 4: Merge Remittance y FBL5N
    # -------------------
    hrc_template = pd.merge(remittance, FBL5N, on="Referencia / Factura", how="left")

    # -------------------
    # Paso 5: Diferencias y registros especiales
    # -------------------
    hrc_template["Diferencia"] = pd.NA
    hrc_template.loc[hrc_template["Tipo de Documento"] == "Factura", "Diferencia"] = (
        hrc_template["importe_FBL5N"] - hrc_template["Importe de factura"]
    )

    diferencias = hrc_template[hrc_template["Diferencia"].notna() & (hrc_template["Diferencia"] != 0)].copy()

    registros_diferencias = pd.DataFrame({
        "Tipo de Documento": "Descuentos no asociados a FC",
        "Referencia / Factura": diferencias["Referencia / Factura"],
        "Importe de factura": diferencias["Diferencia"],
        "Pago Neto": "",
        "Descuento": diferencias["Diferencia"].apply(lambda x: "MENORES VALORES" if -20000 < x < 20000 else "A definir"),
        "Motivo del descuento": diferencias["Diferencia"].apply(lambda x: "WOB" if x < 0 else "384")
    })
    hrc_template = pd.concat([hrc_template, registros_diferencias], ignore_index=True)

    # -------------------
    # Paso 6: Comentarios y Pago Neto
    # -------------------
    hrc_template["Comentarios"] = hrc_template["Descuento"].fillna("") + " " + hrc_template["Referencia / Factura"].fillna("")
    hrc_template["Pago Neto"] = hrc_template["Importe de factura"]

    columnas_finales = [
        "Tipo de Documento",
        "Referencia / Factura",
        "Importe de factura",
        "Descuento",
        "Motivo del descuento",
        "Pago Neto",
        "Comentarios"
    ]
    hrc_template = hrc_template[columnas_finales]

    # -------------------
    # Paso 7: Datos adicionales del Excel
    # -------------------
    # Orden de pago
    wb_rem = load_workbook(remittance_file, data_only=True)
    ws_rem = wb_rem.active
    celda = ws_rem["C10"].value
    numero_orden = celda.split("Orden de Pago:")[1].strip() if celda and "Orden de Pago:" in celda else ""

    # Cliente
    fbl5n_cliente = pd.read_excel(fbl5n_file, usecols=["Customer", "Name 1"], nrows=1)
    id_cliente = fbl5n_cliente["Customer"].iloc[0]
    nombre_cliente = fbl5n_cliente["Name 1"].iloc[0]

    # Fecha y valor banco
    wb_fbl3n = load_workbook(fbl3n_file, data_only=True)
    ws_fbl3n = wb_fbl3n.active
    fecha_original = ws_fbl3n["K8"].value
    fecha_dt = datetime.strptime(fecha_original, "%d.%m.%Y")
    fecha_pago = fecha_dt.strftime("%-m/%-d/%y")
    importe_FBL3N = ws_fbl3n["N8"].value
    importe_FBL3N = abs(float(importe_FBL3N.replace(".", "").replace(",", ".")))

    total_pago_neto = hrc_template["Pago Neto"].sum()
    diferencia = total_pago_neto - importe_FBL3N

    ruta_salida = "Template_HRC.xlsx"
    hrc_template.to_excel(ruta_salida, index=False, sheet_name="Template", startrow=17, startcol=2)

    # -------------------
    # Paso 8: Formatos con openpyxl
    # -------------------
    wb = load_workbook(ruta_salida)
    ws = wb["Template"]

    ws["C2"] = "Desglose de Pago"
    ws["C4"] = "CAMPOS NO EDITABLES"
    ws["G2"] = "REFERENCIA DE PAGO"
    ws["H2"] = numero_orden
    ws["C6"], ws["C8"] = "Cliente", "Codigo de Cliente"
    ws["D6"], ws["D8"] = nombre_cliente, id_cliente
    ws["C11"], ws["D11"], ws["E11"], ws["F11"] = "Referencia", "Fecha", "Método de Pago", "Valor"
    ws["C12"], ws["D12"], ws["E12"], ws["F12"] = numero_orden, fecha_pago, "Transferencia", importe_FBL3N
    ws["F6"], ws["G6"], ws["F7"], ws["G7"], ws["F8"], ws["G8"] = "TOTAL s/ BANCOS", importe_FBL3N, "TOTAL s/ DETALLE", total_pago_neto, "DIFERENCIA", -diferencia

    # Formatos numéricos
    for cell in ["G6", "G7", "G8", "F12"]:
        ws[cell].number_format = '#,##0.00'

    # Formato tabla principal
    num_cols = ["Importe de factura", "Pago Neto"]
    for col in num_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            ws.cell(row=row, column=col_idx).number_format = '#,##0.00'

    str_cols = ["Tipo de Documento", "Referencia / Factura", "Descuento", "Motivo del descuento", "Comentarios"]
    for col in str_cols:
        col_idx = hrc_template.columns.get_loc(col) + 3
        for row in range(18, 18 + len(hrc_template) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = '@'
            if col != "Comentarios":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(ruta_salida)
    print(f"Archivo exportado correctamente con formato: {ruta_salida}")

    return hrc_template, ruta_salida
